# app/data_layer/IO/excel/excel_exporter.py

from typing import TYPE_CHECKING, List, Optional

from openpyxl.utils import get_column_letter

from .base_objects import Events, IFormatter, SpreadsheetExporter

if TYPE_CHECKING:
    from .excel_objects import (ExcelChart, ExcelSheet, ExcelTable, Position,
                                SheetLayoutModel)


class ExcelExporter(SpreadsheetExporter):
    def __init__(self, sheet_generator: Optional['ExcelSheetGenerator'] = None):
        """Handles creating a workbook from excel sheets and exporting to excel file"""
        super().__init__()
        self.map_table_to_chart = {}
        self.sheet_generator = sheet_generator or ExcelSheetGenerator()

    def create_sheet(self, headers, rows, labels=None, sheet_name="Sheet1"):
        excel_sheet = ExcelSheet(headers, rows, labels, sheet_name)
        return excel_sheet
    
    def export_to_spreadsheet(self, excel_worksheets: List['ExcelSheet'], output_file_name: str):
        self.create_progress_bar(len(excel_worksheets))

        for i, excel_sheet in enumerate(excel_worksheets):
            active_worksheet = self.workbook.create_sheet(title=excel_sheet.worksheet_name)
            self.sheet_generator.generate(excel_sheet, active_worksheet)

            self.events.emit("on_progress", (i+1) / len(excel_worksheets) * 100)
            self.update_progress_bar(1)

        if self.progress_bar:
            self.progress_bar.close()

        self.workbook.save(output_file_name)

class ExcelSheetGenerator:
    """Class to generate Excel sheets based on a layout model."""

    def generate(self, excel_sheet: 'ExcelSheet', worksheet):
        layout_model = excel_sheet.layout_model
        self.generate_tables(layout_model.data_tables, worksheet)
        self.generate_charts(layout_model.data_charts, worksheet)

    def generate_tables(self, tables: List['ExcelTable'], worksheet):
        for table in tables:
            row_start, col_start, _, _ = table.position
            col_letters = [get_column_letter(i) for i in range(col_start, col_start + len(table.headers))]

            # Writing headers
            for col_idx, header in enumerate(table.headers):
                worksheet[f"{col_letters[col_idx]}{row_start}"] = header

            # Writing data rows
            for row_idx, row_data in enumerate(table.data.values, start=row_start + 1):  # +1 to skip header row
                for col_idx, cell_value in enumerate(row_data):
                    worksheet[f"{col_letters[col_idx]}{row_idx}"] = cell_value

    def generate_charts(self, charts: List['ExcelChart'], worksheet):
        pass  # Implement chart population logic


# Abstract class for formatting Excel worksheets  
class ExcelFormatter(IFormatter):
    def apply_table_formatting(self, excel_sheet:' ExcelSheet'):
        raise NotImplementedError()
