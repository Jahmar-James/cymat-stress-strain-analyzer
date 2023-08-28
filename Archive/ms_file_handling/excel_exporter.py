# excel_exporter.py
import cProfile
import tkinter as tk

import pandas as pd
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle, Font, PatternFill
import subprocess

# Excel Sheet Names
SELECTED_SPECIMEN = 'Selected Specimens'
SPECIMENS_OVERLAYED ='Specimens Overlay'
AVERAGE_OF_SELECTED=  'Average Curve'
RAW_DATA = 'Raw Data'
PROCESSED_DATA = 'Processed Data'
SUMMARY = 'Summary'
AVERAGE_CHART = "Average Chart"

class ExcelExporter:
    """
    This class handles the exporting of data to an Excel file.

    Attributes:
        app: A reference to the application that uses this class.
        export_in_progress: A boolean flag indicating whether an export is currently in progress.
    """
    def __init__(self, app):
        self.app = app
    
    @property
    def selected_specimens(self):
        selected_indices = self.app.widget_manager.specimen_listbox.curselection()
        return [self.app.variables.specimens[index] for index in selected_indices]
    
    def profile_export_average_to_excel(self,selected_indices, file_path):

        # cProfile.runctx('self.export_data_to_excel(selected_indices, file_path)', globals(), locals(), 'export_data_to_excel.profile')
        # Run the function with profiling and save the result
        profiler = cProfile.Profile()
        profiler.runcall(self.export_data_to_excel, selected_indices, file_path)
        
        # Dump the profiling results to a file
        profile_file = 'export_data_to_excel.profile'
        profiler.dump_stats(profile_file)

        subprocess.run(["snakeviz", profile_file])

    # Main control flow function 
    def export_data_to_excel(self, selected_indices, file_path):
        """
        Exports the data of the selected specimens to an Excel file.

        Args:
            selected_indices: Indices of the specimens to export.
            file_path: The path to the Excel file to which the data is exported.
        """
        print("export_data_to_excel")
        def create_charts(writer, data_dfs, average_df):
            # Create combined chart for selected specimens
            selected_specimens_ws = writer.sheets[SELECTED_SPECIMEN ]
            
            chart1 = ScatterChart()
            chart1.title = "Stress-Strain Curve "

            chart2 = ScatterChart()
            chart2.title = "Stress-Shifted Strain Curve "

            chart3 = ScatterChart()
            chart3.title = "Force-Shifted Displacement Curve "


            for idx, df in enumerate(data_dfs):
                specimen = self.selected_specimens[idx]
                stress_col = idx * (len(df.columns) + 1) + 4
                strain_col = stress_col + 1
                shifted_strain_col = stress_col + 2
                force_col = stress_col -2
                shifted_disp_col = stress_col + 3
                row = 5

                # Stress-Strain Curve
                x_data = Reference(selected_specimens_ws, min_col=strain_col, min_row=row, max_col=strain_col, max_row=len(df) + 5)
                y_data = Reference(selected_specimens_ws, min_col=stress_col, min_row=row, max_col=stress_col, max_row=len(df) + 5)
                series = Series(values=y_data, xvalues=x_data, title=f"Specimen {specimen.name}")
                chart1.series.append(series)

                # Stress-Shifted Strain Curve
                x_data_2 = Reference(selected_specimens_ws, min_col=shifted_strain_col, min_row=row, max_col=shifted_strain_col, max_row=len(df) + 5)
                y_data_2 = Reference(selected_specimens_ws, min_col=stress_col, min_row=row, max_col=stress_col, max_row=len(df) + 5)
                series_2 = Series(values=y_data_2, xvalues=x_data_2, title=f"Specimen {specimen.name}")
                chart2.series.append(series_2)

                # Force-Shifted Displacement Curve
                x_data_3 = Reference(selected_specimens_ws, min_col=shifted_disp_col, min_row=row, max_col=shifted_disp_col, max_row=len(df) + 5)
                y_data_3 = Reference(selected_specimens_ws, min_col=force_col, min_row=row, max_col=force_col, max_row=len(df) + 5)
                series_3 = Series(values=y_data_3, xvalues=x_data_3, title=f"Specimen {specimen.name}")
                chart3.series.append(series_3)

            chart1.x_axis.title = "Strain"
            chart1.y_axis.title = "Stress (MPa)"

            chart2.x_axis.title = "Shifted Strain"
            chart2.y_axis.title = "Stress (MPa)"

            chart3.x_axis.title = "Shifted Displacement (mm)"
            chart3.y_axis.title = "Force (N)"
            
            # Create a new sheet for the combined chart of selected specimens
            combined_chart_ws = writer.book.create_sheet(SPECIMENS_OVERLAYED)
            combined_chart_ws.add_chart(chart1, "A" + str( 1))
            combined_chart_ws.add_chart(chart2, "M" + str(1))
            combined_chart_ws.add_chart(chart3, "W" + str(1))

            # Create chart for the average curve
            average_ws = writer.sheets[ AVERAGE_OF_SELECTED]
            chart4 = ScatterChart()
            chart4.title = "Stress-Strain Curve - Average"
            x_data = Reference(average_ws, min_col=3, min_row=2, max_col=3, max_row=len(average_df) + 1)
            y_data = Reference(average_ws, min_col=4, min_row=2, max_col=4, max_row=len(average_df) + 1)
            series = Series(values=y_data, xvalues=x_data, title="Average")
            chart4.series.append(series)

            chart4.x_axis.title = "Strain"
            chart4.y_axis.title = "Stress (MPa)"
            
            # Create a new sheet for the average curve chart
            average_chart_ws = writer.book.create_sheet(AVERAGE_CHART)
            average_chart_ws.add_chart(chart4, "A1")

        def apply_formatting(worksheet,apply_pattern_fill=True):
            column_widths = {}

            for i, row in enumerate(worksheet.iter_rows(), start=1):
                for cell in row:
                    # Determine column widths
                    column_widths[cell.column] = max(column_widths.get(cell.column, 0), len(str(cell.value)))

                    if i == 1:
                        cell.style = 'header_font'
                    elif apply_pattern_fill and i % 2 == 0:  # Apply alternating row fill
                        cell.style = 'fill'
                    elif isinstance(cell.value, (int, float)):
                        cell.style = 'number_format'

            # Set column widths
            for column, width in column_widths.items():
                worksheet.column_dimensions[get_column_letter(column)].width = width + 2
            # arial_font = Font(name='Arial', size=8)
            # header_font = Font(name='Arial', size=10, bold=True)
            # fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # number_format = '#,##0.000' 

            # column_widths = {}
            
            # for i, row in enumerate(worksheet.iter_rows(), start=1):
            #     for cell in row:
            #         # Apply font to all cells
            #         cell.font = arial_font
                    
            #         # Determine column widths
            #         column_widths[cell.column] = max(column_widths.get(cell.column, 0), len(str(cell.value)))
                    
            #         # Apply formatting to header row
            #         if i == 1:
            #             cell.font = header_font
            #         elif apply_pattern_fill and i % 2 == 0:   # Apply alternating row fill
            #             cell.fill = fill
                    
            #         if isinstance(cell.value, (int, float))and sheet_name not in [SPECIMENS_OVERLAYED, AVERAGE_CHART]:
            #             cell.number_format = number_format
            
            # # Set column widths
            # for column, width in column_widths.items():
            #     worksheet.column_dimensions[get_column_letter(column)].width = width + 2

        def add_summary_sheet(writer):
            summary_dfs = []

            summary_stats_df = self.app.data_handler.summary_statistics()

            summary_dfs.append(summary_stats_df)
            
            average_summary = self.app.variables.average_of_specimens.describe().transpose()
            summary_dfs.append(average_summary)
            
            for df in self.data_dfs:
                summary = df.describe().transpose()
                summary_dfs.append(summary)
            
            row_offset = 0
            summary_ws = writer.book.create_sheet(SUMMARY)

            specimen = None
            for i, summary_df in enumerate(summary_dfs):
                if i == 0:
                    specimen_name = "Summary Statistics"
                elif i == 1:
                    specimen_name = "Average"
                else:
                    specimen = self.selected_specimens[i - 2]
                    specimen_name = specimen.name if specimen else "Unknown Specimen"

                if specimen:
                    if specimen.processed_hysteresis_data is None and specimen.IYS:
                        yield_stress, yield_strain = specimen.IYS
                        density_iys_text = f" ({specimen.density:.2f} g/cc,IYS: {yield_stress:.2f} MPa, {yield_strain:.2f} mm)"
                    else:
                        density_iys_text = f" ({specimen.density:.2f} g/cc"
                    
                else: density_iys_text = ""
                summary_ws.cell(row=row_offset + 1, column=1, value=specimen_name + density_iys_text).font = Font(bold=True)
                row_offset += 1
                summary_df.to_excel(writer, sheet_name=SUMMARY, index=True, startrow=row_offset)
                row_offset += len(summary_df) + 1

            
            # apply_formatting(summary_ws)

        properties_dfs, data_dfs = self.prepare_data(selected_indices)
        self.properties_dfs = properties_dfs
        self.data_dfs =data_dfs

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                self.set_workbook_style(writer.book)
                self.write_dfs_to_excel(properties_dfs, data_dfs, writer)
                # self.app.variables.average_of_specimens.to_excel(writer, sheet_name= AVERAGE_OF_SELECTED, index=False)
                avg_start_row = self.write_average_of_specimens(writer, AVERAGE_OF_SELECTED)
                self.create_table(writer,  AVERAGE_OF_SELECTED, avg_start_row, 0, len(self.app.variables.average_of_specimens.index), len(self.app.variables.average_of_specimens.columns))
                self.write_raw_data_to_excel(writer)
                self.write_processed_data_to_excel(writer) 

                for sheet_name in [SELECTED_SPECIMEN ,  AVERAGE_OF_SELECTED]:
                    ws = writer.sheets[sheet_name]
                    apply_formatting(ws,apply_pattern_fill = False )
                    ws.freeze_panes = ws.cell(row=5 if sheet_name == SELECTED_SPECIMEN  else 2, column=1)
                add_summary_sheet(writer)
                create_charts(writer, data_dfs, self.app.variables.average_of_specimens)
                self.app.variables.export_in_progress = False
                tk.messagebox.showinfo("Success", f"Data successfully exported to {file_path}")

        except Exception as e:
            tk.messagebox.showerror("Error", f"An error occurred while exporting the data: {e}")
            self.app.variables.export_in_progress = False
   
    def set_workbook_style(self, workbook):
        self.wb = workbook

        arial_font = NamedStyle(name="arial_font")
        arial_font.font = Font(name='Arial', size=8)
        self.wb.add_named_style(arial_font)

        header_font = NamedStyle(name="header_font")
        header_font.font = Font(name='Arial', size=10, bold=True)
        self.wb.add_named_style(header_font)

        fill = NamedStyle(name="fill")
        fill.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        self.wb.add_named_style(fill)

        number_format = NamedStyle(name="number_format")
        number_format.number_format = '#,##0.000'
        self.wb.add_named_style(number_format)

    # Procoessing and creation functions
    def prepare_data(self, selected_indices):
        properties_dfs = []
        data_dfs = []

        for index in selected_indices:
            specimen = self.app.variables.specimens[index]
            properties_df = self.get_properties_df(specimen)
            properties_dfs.append(properties_df)
            data_dfs.append(specimen.shifted_data)

        return properties_dfs, data_dfs

    def get_properties_df(self, specimen):
        # Create common part of DataFrame
        df = pd.DataFrame({
            'Specimen Name': [specimen.name],
            'Density (g/cc)': [specimen.density],
            'Cross-sectional Area (mm^2)': [specimen.cross_sectional_area],
            'Original Length (mm)': [specimen.original_length]
        })

        # Append columns based on conditions
        if specimen.processed_hysteresis_data is not None and specimen.graph_manager.compressive_proof_strength:
            ps_strain, ps_stress = specimen.graph_manager.compressive_proof_strength
            df['Proof Strain'] = [ps_strain]
            df['Proof Stress'] = [ps_stress]
            df['Elastic Gradient'] = [specimen.data_manager.modulus]
        elif specimen.IYS:
            iys_strain, iys_stress = specimen.IYS
            df['IYS Strain'] = [iys_strain]
            df['IYS Stress'] = [iys_stress]
            df['Youngs Modulus'] = [specimen.youngs_modulus]

        return df
    
    def _write_dfs_to_excel(self, properties_dfs, data_dfs, writer, start_row=0, start_col=0):
        """
        Writes properties and data dataframes to an Excel file.

        Args:
            properties_dfs: A list of properties dataframes to write.
            data_dfs: A list of data dataframes to write.
            writer: The ExcelWriter to use for writing.
            start_row: The row at which to start writing (default is 0).
            start_col: The column at which to start writing (default is 0).
        """

        # Compute the max length of columns for all dataframes only once
        max_len = max(len(df.columns) for df in properties_dfs + data_dfs)

        # Add empty columns between dataframes in properties_dfs and data_dfs
        properties_dfs_with_space = [df for df in properties_dfs for _ in range(2)]
        data_dfs_with_space = [df for df in data_dfs for _ in range(2)]

        # Concatenate all properties dataframes and write to Excel
        properties_df = pd.concat(properties_dfs_with_space, keys=range(len(properties_dfs_with_space)), axis=1)
        properties_df.columns = properties_df.columns.get_level_values(1)  # Use original column names
        properties_df.to_excel(writer, sheet_name=SELECTED_SPECIMEN, index=False, startrow=start_row, startcol=start_col)

        # Create tables for each original dataframe in properties_dfs
        start_col_temp = start_col
        for df in properties_dfs:
            self.create_table(writer, SELECTED_SPECIMEN, start_row, start_col_temp, len(df.index), len(df.columns))
            start_col_temp += max_len + 1

        start_col = 0  # Reset start column

        # Concatenate all data dataframes and write to Excel
        data_df = pd.concat(data_dfs_with_space, keys=range(len(data_dfs_with_space)), axis=1)
        data_df.columns = data_df.columns.get_level_values(1)  # Use original column names
        data_df.to_excel(writer, sheet_name=SELECTED_SPECIMEN, index=False, startrow=start_row + 3, startcol=start_col)

        # Create tables for each original dataframe in data_dfs
        start_col_temp = start_col
        for df in data_dfs:
            self.create_table(writer, SELECTED_SPECIMEN, start_row + 3, start_col_temp, len(df.index), len(df.columns))
            start_col_temp += max_len + 1


    def write_dfs_to_excel(self, properties_dfs, data_dfs, writer, start_row=0, start_col=0):
        """
        Writes properties and data dataframes to an Excel file.

        Args:
            properties_dfs: A list of properties dataframes to write.
            data_dfs: A list of data dataframes to write.
            writer: The ExcelWriter to use for writing.
            start_row: The row at which to start writing (default is 0).
            start_col: The column at which to start writing (default is 0).
        """

         # Compute the max length of columns for all dataframes only once
        max_len = max(len(df.columns) for df in properties_dfs + data_dfs)

        # Write properties tables
        for df in properties_dfs:
            df.columns = df.columns.astype(str)
            df.to_excel(writer, sheet_name=SELECTED_SPECIMEN , index=False, startrow=start_row, startcol=start_col)
            self.create_table(writer, SELECTED_SPECIMEN , start_row, start_col, len(df.index), len(df.columns))
            start_col += max_len + 1

        start_col = 0
        
        # Write data tables
        for df in data_dfs:
            df.to_excel(writer, sheet_name=SELECTED_SPECIMEN , index=False, startrow=start_row + 3, startcol=start_col)
            self.create_table(writer, SELECTED_SPECIMEN , start_row + 3, start_col, len(df.index), len(df.columns))
            start_col +=  max_len + 1

    def write_average_of_specimens(self, writer, sheet_name):
        # Create a new DataFrame with your descriptions
        descriptions = pd.DataFrame({'Description': ['Average Data', 'description 2', 'description 3']})
        start_row = len(descriptions)
        descriptions.to_excel(writer, sheet_name=sheet_name, index=False)

        # Write the average_of_specimens DataFrame to Excel, starting after the descriptions
        self.app.variables.average_of_specimens.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)

        # Write the average_of_specimens_hysteresis DataFrame to the same sheet, leaving a column of space in between
        if self.app.variables.average_of_specimens_hysteresis is not None:
            self.app.variables.average_of_specimens_hysteresis.to_excel(writer, sheet_name=sheet_name, startcol=self.app.variables.average_of_specimens.shape[1] + 1,startrow=start_row, index=False)
        return start_row



    def write_raw_data_to_excel(self, writer):
        raw_data_dfs = [specimen.data for specimen in self.selected_specimens]
        for idx, df in enumerate(raw_data_dfs):
            df.to_excel(writer, sheet_name=RAW_DATA, index=False, startrow=1, startcol=idx * (len(df.columns) + 1))
            ws = writer.sheets[RAW_DATA]
            ws.cell(row=1, column=idx * (len(df.columns) + 1) + 1, value=self.selected_specimens[idx].name).font = Font(bold=True)

    def write_processed_data_to_excel(self, writer):
        PROCESSED_DATA_HYSTERESIS = 'Hysteresis data'  # the sheet name for hysteresis data

        for idx, specimen in enumerate(self.selected_specimens):
            startcol = idx * (len(specimen.processed_data.columns) + 1)
            
            # Write processed data
            specimen.processed_data.to_excel(writer, sheet_name=PROCESSED_DATA, index=False, startrow=1, startcol=startcol)
            ws = writer.sheets[PROCESSED_DATA]
            ws.cell(row=1, column=startcol + 1, value=specimen.name).font = Font(bold=True)
            
            # If available, write hysteresis data into a new sheet
            if  specimen.processed_hysteresis_data is not None and not specimen.processed_hysteresis_data.empty:
                specimen.processed_hysteresis_data.to_excel(writer, sheet_name=PROCESSED_DATA_HYSTERESIS, index=False, startrow=1, startcol=startcol)
                ws_hysteresis = writer.sheets[PROCESSED_DATA_HYSTERESIS]
                ws_hysteresis.cell(row=1, column=startcol + 1, value=specimen.name).font = Font(bold=True)
      

    def create_table(self, writer, sheet_name, start_row, start_col, row_count, col_count):
        """
        Creates a table in an Excel file.

        Args:
            writer: The ExcelWriter to use for writing.
            sheet_name: The name of the sheet in which to create the table.
            start_row: The row at which to start the table.
            start_col: The column at which to start the table.
            row_count: The number of rows in the table.
            col_count: The number of columns in the table.
        """
        def get_used_table_names(workbook):
            """Get a list of all table names used in the workbook."""
            return [
                t.displayName
                for sheet in workbook
                for t in sheet._tables.values()
            ]

        worksheet = writer.sheets[sheet_name]
        data_range = f"{get_column_letter(start_col + 1)}{start_row + 1}:{get_column_letter(start_col + col_count)}{start_row + row_count + 1}"
        base_table_name = f"{sheet_name.replace(' ', '_')}_Table"
        table_name = base_table_name
        idx = 1
        used_table_names = get_used_table_names(writer.book)
        while table_name in used_table_names:
            table_name = f"{base_table_name}_{idx}"
            idx += 1

        table = Table(displayName=table_name, ref=data_range)
        style = TableStyleInfo(name="TableStyleLight11", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)





