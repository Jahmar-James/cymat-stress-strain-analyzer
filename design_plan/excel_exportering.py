from abc import ABC, abstractmethod
from collections import namedtuple
from contextlib import contextmanager
from dataclasses import dataclass
from typing import List, Union

import numpy as np
import pandas as pd
import tqdm
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Color, Font, PatternFill,
                             Protection, Side)
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, root_validator

# Class to represent the position of a table or chart within a worksheet.
Position = namedtuple("Position", ["start_row", "start_col", "end_row", "end_col"])

@dataclass
class ExcelTable:
    """
    Class to represent a table within an Excel worksheet
    """
   
    headers: List[str]
    data: pd.DataFrame
    position: Position

    
    @classmethod
    def from_dataframe(cls, df: pd.DataFrame, position: Position):
        # if only the "start_row", "start_col" are provided, then end row and end col are calculated
        # from the length of the dataframe
        headers = list(df.columns)
        data = df
        return cls(headers=headers, data=data, position=position)

@dataclass
class ExcelChart:
    data_range: Position
    position: Position

class SheetLayoutModel(BaseModel):
    """Class to represent the layout of a single worksheet with tables and charts"""
    data_tables: List[ExcelTable] = Field(default_factory=list)
    data_charts: List[ExcelChart] = Field(default_factory=list)

    @root_validator
    def validate_no_overlap(cls, layout_elements):
        all_positions = [table.position for table in layout_elements.get('data_tables', [])] + \
                        [chart.position for chart in layout_elements.get('data_charts', [])]

        if len(all_positions) != len(set(all_positions)):
            raise ValueError('Overlap detected.')
        return layout_elements


class ExcelSheet:
    """Class to represent a single worksheet with tables and charts"""
    def __init__(self, worksheet_name: str = "Sheet1"):
        self.worksheet_name = worksheet_name
        self.layout_model = SheetLayoutModel()

    def add_table(self, table: ExcelTable):
        self.layout_model.data_tables.append(table)
        self.layout_model = SheetLayoutModel(**self.layout_model.dict())

    def add_chart(self, chart: ExcelChart):
        self.layout_model.data_charts.append(chart)
        self.layout_model = SheetLayoutModel(**self.layout_model.dict())

    def generate_tables_for_worksheet(self, worksheet):
        for table in self.layout_model.data_tables:
            row_start, col_start, _, _ = table.position
            for row_idx, row_data in enumerate(table.data.iterrows(), start=row_start):
                for col_idx, cell_value in enumerate(row_data[1], start=col_start):
                    col_letter = get_column_letter(col_idx)
                    worksheet[f"{col_letter}{row_idx}"] = cell_value

    def generate_charts_for_worksheet(self, worksheet):
        # TODO: Implement chart population logic
        pass

class ExcelSheetGenerator:
    """Class to generate Excel sheets based on a layout model."""

    def generate(self, excel_sheet: ExcelSheet, worksheet):
        layout_model = excel_sheet.layout_model
        self.generate_tables(layout_model.data_tables, worksheet)
        self.generate_charts(layout_model.data_charts, worksheet)

    def generate_tables(self, tables: List[ExcelTable], worksheet):
        for table in tables:
            row_start, col_start, _, _ = table.position
            col_letters = [get_column_letter(i) for i in range(col_start, col_start + len(table.headers))]

            # Writing headers
            for col_idx, header in enumerate(table.headers):
                worksheet[f"{col_letters[col_idx]}{row_start}"] = header

            # Writing data rows
            for row_idx, row_data in enumerate(table.data.values, start=row_start + 1):  # +1 to skip header row
                for col_idx, cell_value in enumerate(row_data):
                    worksheet[f"{col_letters[col_idx]}{row_idx}"] = cell_value

    def generate_charts(self, charts: List[ExcelChart], worksheet):
        pass  # Implement chart population logic

class IFormatter(ABC):
    @abstractmethod
    def apply_table_formatting(self, sheet):
        pass
    
    @abstractmethod
    def set_font_styles(self, sheet, font_options):
        pass
    
    @abstractmethod
    def set_color_schemes(self, sheet, color_options):
        pass

# Abstract class for formatting Excel worksheets  
class ExcelFormatter(IFormatter):
    def apply_table_formatting(self, excel_sheet: ExcelSheet):
        raise NotImplementedError()

class Events:
    def __init__(self):
        self.handlers = {}
    
    def on(self, event_name, handler):
        if event_name not in self.handlers:
            self.handlers[event_name] = []
        self.handlers[event_name].append(handler)
    
    def emit(self, event_name, *args, **kwargs):
        if event_name in self.handlers:
            for handler in self.handlers[event_name]:
                handler(*args, **kwargs)

class SpreadsheetExporter(ABC):
    def __init__(self, workbook = None):
        self.workbook = workbook or Workbook()
        self.sheets = []
        self.file_name = None
        self.events = Events()
        self.progress_bar = None
    
    def create_progress_bar(self, total):
        self.progress_bar = tqdm.tqdm(total=total, desc="Exporting Worksheets", position=0, leave=True)

    def update_progress_bar(self, value):
        if self.progress_bar:
            self.progress_bar.update(value)

    @contextmanager
    def exporting_process(self, sheets: List[ExcelSheet], file_name: str):
        self.create_progress_bar(len(sheets))
        try:
            yield  # Control is transferred to the caller's block
            self.export_to_spreadsheet(sheets, file_name)
        finally:
            if self.progress_bar:
                self.progress_bar.close()

    @abstractmethod       
    def export_to_spreadsheet(self, sheets: List[ExcelSheet], file_name: str) -> None:
        pass
        
# Class to export Excel worksheets into a workbook
class ExcelExporter(SpreadsheetExporter):
    def __init__(self, sheet_generator: ExcelSheetGenerator = None):
        """Handles creating a workbook from excel sheets and exporting to excel file"""
        super().__init__()
        self.map_table_to_chart = {}
        self.sheet_generator = sheet_generator or ExcelSheetGenerator()

    def create_sheet(self, headers, rows, labels=None, sheet_name="Sheet1"):
        excel_sheet = ExcelSheet(headers, rows, labels, sheet_name)
        return excel_sheet
    
    def export_to_spreadsheet(self, excel_worksheets: List[ExcelSheet], output_file_name: str):
        self.create_progress_bar(len(excel_worksheets))

        for i, excel_sheet in enumerate(excel_worksheets):
            active_worksheet = self.workbook.create_sheet(title=excel_sheet.worksheet_name)
            self.sheet_generator.generate(excel_sheet, active_worksheet)

            self.events.emit("on_progress", (i+1) / len(excel_worksheets) * 100)
            self.update_progress_bar(1)

        if self.progress_bar:
            self.progress_bar.close()

        self.workbook.save(output_file_name)


class ExcelExportService:
    """Handles Excel sheet formatting and exporting"""
    def __init__(self, formatter: ExcelFormatter = None, excel_exporter: ExcelExporter = None):
        self.formatter = formatter or ExcelFormatter()
        self.exporter = excel_exporter or ExcelExporter()

    def format_sheet(self, excel_sheet: ExcelSheet, font_options=None, color_options=None):
        self.formatter.apply_table_formatting(excel_sheet)
        if font_options:
            self.formatter.set_font_styles(excel_sheet, font_options)
        if color_options:
            self.formatter.set_color_schemes(excel_sheet, color_options)

    def export_multiple_to_excel(self, excel_sheets: List[ExcelSheet], file_name: str):
        with self.exporter.exporting_process(excel_sheets, file_name):
            pass

# Example usage
formatter = ExcelFormatter()
exporter = ExcelExporter()
export_service = ExcelExportService(formatter, exporter)

sheet_data = pd.DataFrame(np.random.randn(5, 4), columns=['A', 'B', 'C', 'D'])
sheet1 = exporter.create_sheet(['A', 'B', 'C', 'D'], sheet_data, sheet_name="MySheet")

font_options = {'font_name': 'Arial', 'font_size': 10}
color_options = {'header': 'blue', 'rows': 'white'}

export_service.format_sheet(sheet1, font_options, color_options)
export_service.export_multiple_to_excel([sheet1], 'example.xlsx')
